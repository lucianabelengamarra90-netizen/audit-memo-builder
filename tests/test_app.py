import io
import unittest

from openpyxl import Workbook, load_workbook
from pypdf import PdfWriter

from app import app, detect_categories, extract_pdf, extract_xlsx


def workbook_bytes(sheets):
    workbook = Workbook()
    first = True
    for name, rows in sheets:
        worksheet = workbook.active if first else workbook.create_sheet()
        first = False
        worksheet.title = name
        for row in rows:
            worksheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


class HallazgoDetectorTests(unittest.TestCase):
    def test_only_matches_complete_singular_or_plural_word(self):
        self.assertEqual(detect_categories("HALLAZGO documentado")[0]["keyword"], "hallazgo")
        self.assertEqual(detect_categories("Dos hallazgos relevantes")[0]["keyword"], "hallazgos")
        for text in ("observación", "riesgo", "diferencia", "hallazgoso", "prehallazgo"):
            self.assertEqual(detect_categories(text), [], text)

    def test_normalization_does_not_change_category_contract(self):
        self.assertEqual(detect_categories("  Hallazgo   crítico  "), [
            {"category": "Hallazgo", "keyword": "hallazgo"}
        ])


class ExcelExtractionTests(unittest.TestCase):
    def test_reads_every_sheet_and_preserves_traceability(self):
        source = workbook_bytes([
            ("Inicio", [["sin coincidencia"], ["Hallazgo de caja"]]),
            ("Final", [["otro dato"], ["Hallazgos de inventario"]]),
        ])
        items, warnings = extract_xlsx(type("Upload", (), {"stream": source})(), "prueba.xlsx")
        self.assertEqual(warnings, [])
        self.assertEqual(len(items), 2)
        self.assertEqual({item["originName"] for item in items}, {"Inicio", "Final"})
        self.assertEqual({item["reference"] for item in items}, {"Fila 2"})
        self.assertTrue(all(item["category"] == "Hallazgo" for item in items))
        self.assertEqual({item["keyword"] for item in items}, {"hallazgo", "hallazgos"})

    def test_does_not_inherit_header_or_column_category(self):
        source = workbook_bytes([
            ("Datos", [
                ["Hallazgos"],
                ["Esta diferencia no repite la palabra buscada"],
                ["Detalle", "Riesgo", "Monto"],
                ["Caso 1", "Alto", 10],
            ]),
        ])
        items, _ = extract_xlsx(type("Upload", (), {"stream": source})(), "prueba.xlsx")
        self.assertEqual(len(items), 1)
        self.assertEqual(items[0]["text"], "Hallazgos")
        self.assertEqual(items[0]["reference"], "Fila 1")

    def test_keeps_more_than_250_literal_matches(self):
        source = workbook_bytes([
            ("Masiva", [[f"Hallazgo {number}"] for number in range(1, 302)]),
        ])
        items, warnings = extract_xlsx(type("Upload", (), {"stream": source})(), "masiva.xlsx")
        self.assertEqual(len(items), 301)
        self.assertEqual(warnings, [])
        self.assertEqual(items[-1]["reference"], "Fila 301")

    def test_scanned_pdf_returns_warning_but_no_non_finding_item(self):
        output = io.BytesIO()
        writer = PdfWriter()
        writer.add_blank_page(width=100, height=100)
        writer.write(output)
        output.seek(0)
        items, warnings = extract_pdf(output, "escaneado.pdf")
        self.assertEqual(items, [])
        self.assertEqual(len(warnings), 1)
        self.assertIn("escaneado.pdf", warnings[0])


class ApiRegressionTests(unittest.TestCase):
    def setUp(self):
        self.client = app.test_client()

    def test_extract_count_contains_only_literal_hallazgos(self):
        response = self.client.post("/extract", data={
            "freeText": "Riesgo alto\nHallazgo confirmado\nObservación abierta\nHallazgos pendientes"
        })
        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(payload["count"], 2)
        self.assertTrue(all(item["category"] == "Hallazgo" for item in payload["items"]))

    def test_scanned_pdf_warning_is_propagated_outside_items(self):
        pdf = io.BytesIO()
        writer = PdfWriter()
        writer.add_blank_page(width=100, height=100)
        writer.write(pdf)
        pdf.seek(0)
        response = self.client.post("/extract", data={
            "files": (pdf, "escaneado.pdf")
        }, content_type="multipart/form-data")
        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(payload["items"], [])
        self.assertEqual(payload["count"], 0)
        self.assertEqual(len(payload["warnings"]), 1)
        self.assertIn("escaneado.pdf", payload["warnings"][0])

    def test_export_excel_keeps_navigation_contract_and_trace_sheet(self):
        item = {
            "category": "Hallazgo", "text": "Hallazgo confirmado", "filename": "base.xlsx",
            "originName": "Datos", "reference": "Fila 7", "keyword": "hallazgo",
            "included": True, "converted": True,
        }
        response = self.client.post("/export-excel", json={"memo": {
            "general": {"title": "Prueba"}, "findings": [], "sources": [], "extracted": [item]
        }})
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.data))
        self.assertEqual(workbook.sheetnames, ["Memo", "Hallazgos", "Fuentes", "Trazabilidad"])
        trace = workbook["Trazabilidad"]
        self.assertEqual(trace["A2"].value, "Hallazgo")
        self.assertEqual(trace["D2"].value, "Datos")
        self.assertEqual(trace["E2"].value, "Fila 7")
        self.assertEqual(trace["F2"].value, "hallazgo")


if __name__ == "__main__":
    unittest.main()
